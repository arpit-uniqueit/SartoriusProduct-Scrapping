
from bs4 import BeautifulSoup
import time
import json
import requests
from lxml import html
import os
from openpyxl import Workbook, load_workbook
import pricedetails

headers = {
    'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
    'accept-language': 'en-IN,en-GB;q=0.9,en-US;q=0.8,en;q=0.7,gu;q=0.6',
    'cache-control': 'max-age=0',
    'priority': 'u=0, i',
    'sec-ch-ua': '"Chromium";v="136", "Google Chrome";v="136", "Not.A/Brand";v="99"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Windows"',
    'sec-fetch-dest': 'document',
    'sec-fetch-mode': 'navigate',
    'sec-fetch-site': 'same-origin',
    'sec-fetch-user': '?1',
    'upgrade-insecure-requests': '1',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/136.0.0.0 Safari/537.36',
    # 'cookie': 'NEXT_LOCALE=en-us; _gcl_au=1.1.418745738.1746937171; _ga=GA1.1.1092704647.1746937171; dd_anonymous_id=81a439fc-eceb-41ee-abd5-c286b11d3518; _fbp=fb.1.1746937171635.777369643318052359; _evga_d46d={%22uuid%22:%22effaa041fe8d1976%22}; _sfid_4dc9={%22anonymousId%22:%22effaa041fe8d1976%22%2C%22consents%22:[{%22consent%22:{%22provider%22:%22OneTrust%22%2C%22purpose%22:%22Personalization%22%2C%22status%22:%22Opt%20In%22}%2C%22lastUpdateTime%22:%222025-05-11T04:19:32.126Z%22%2C%22lastSentTime%22:%222025-05-11T04:19:32.167Z%22}]}; _mibhv=anon-1746937172199-9295845162_7395; OptanonAlertBoxClosed=2025-05-11T04:19:32.904Z; NEXT_COUNTRY=US; bm_ss=ab8e18ef4e; bm_mi=1C2AB5EA3A4D704DECDA755F88FBA922~YAAQPGw/F4FJLYeWAQAA+KT4xxsFNor77lpZ1AWcY8zt3zgT7/okBDYJ9ujzZYXYAAodegX9bthmYfeNQFqps3/G8BtuDnUBev5k79HOn1i/j8cNT+fm6ol/ambbcwTz05SbdTK/GCL8vOpTcSYPlEGa/cV46DUyv3AIHi7wiWmDm7Mlhm9v0Juj5lWDcnRpjVSMV9QlcDDQfKs3WRJrxF1JN01WV6On2dl02GeSa1x32+q/uj8ZKnuFllYW6rLbsv9zZLyB/JiuA8CAHjHslgka+J+vNvg9KHyp1qT51AuYsb9X/hggL5+g2kAFfN8+6+Th4m2aqCpCSxDEOfyUfn1bs6g0zRh4AAHxfetQcnACpvbCKme5agB2GEhH/t2V4TXMeFCmvY+cXw==~1; ak_bmsc=2DA8C46D9C81644A7CFAE80BC7133308~000000000000000000000000000000~YAAQPGw/FwBKLYeWAQAA3bD4xxsYp4sF0vTYCX6oPYRDh+xgeuGBFuDWCFD/fhn0uf79FO9kKJGOwh5A8JJrG6LMF79Ft/rU0Gh/Y9zcjYXdwMSnh0YyBtiPe9PqaX2ML3bnsHmYjZXLxDyHgs2xxf28vm+dAchQhCFKGyQTKyFZHsc3w1aeffTWSAbLZtOwf7DXvB9b1/HV5CHJBB7P7Z0YIu4aL6lnKKjgaiDX/R1onCKfslNXFx6Id/o//wTiePd7j81P9D3dm3Om7NqABXRv4DwJawB2hQOkqteB7UilRga7e6bC4xOpL10I2WNrVvdUXYnI1S7oLpysa+2G9tMJZcmwb5XpYl7C27obClkq4X8wnpXa0JbtXG8JKcJWISEJ6Vl+rKknZtriCQlClNRAyPHT8ths8DvmmWDwYgmKwjrkFS8ssP7u43jJBrnooqJF7M6sYpbMxTGTmVpky6ayjYIkufyW8dU/d0FZDq7APdlEoKKGX2G1xnd+EGuziQqOihoZ6s43LoeO9bVX2mBIvPG7jss6hL8xWA8r7eY=; bm_so=5BCEDC24BB519919C5B39E3615983A9F9FE94C1449427DCAD0801485020739D9~YAAQPGw/F0t0LYeWAQAASsv8xwOXyhGJElDa++rnq11hVJp+MPvPuWkEeQu5vniwlSLSf1+RDl2rNHpvM7P21/I8cWRGYnZixGz8UfwGvaeswUayUwPJ09DOWzEmtdpfI2+Y/XRrbTA3r/KhuC21MN0bl2iHRnkRB5SNT9nvcK6dIuf23t8naiQp4dVFn5LggHTww0/+Wt5jW+c96HoqhB5WzsEUTvFy02JSqnNuSxE/HXsRK6bfDNVVd8QK2kif+bYkGR+xiDGAqOWmpIN61G2kvaw2Oje+8t/LC+9vM5XbSipZvYN0Z9qLLPaRP77X2Nnvbr2D37q2cK8yEVFPK1inbTr/bU9p9OshHtVre3Osaefz7r4bnFs+sRo858ArHAgGVfuebLSYKyb/DWgV0SL3J6BaIL6WCWlbVSPf6qsoR+qXtRgTt6MnJQwAaVqjWBOtgs6oQTCxsbR+rY/G; OptanonConsent=isGpcEnabled=0&datestamp=Tue+May+13+2025+10%3A22%3A39+GMT%2B0530+(India+Standard+Time)&version=202503.2.0&browserGpcFlag=0&isIABGlobal=false&hosts=&consentId=5882345f-9f01-4672-8837-8392047f17d9&interactionCount=1&isAnonUser=1&landingPath=NotLandingPage&groups=C0001%3A1%2CC0002%3A1%2CC0003%3A1%2CC0004%3A1%2CC0005%3A1&intType=1&geolocation=IN%3BGJ&AwaitingReconsent=false; sa-user-id=s%253A0-e4ac9924-5454-4ecd-429e-fff68654b505.x2upGtVT7YDU%252B7PSkRBmdwr10De6R1JzQD2%252FHECt7f4; sa-user-id-v2=s%253A5KyZJFRUTs1Cnv_2hlS1BQ.w0X7WOMCy0ImHHuUS4KnrYFtYP3kOb%252BI8uOfkgD1wnY; sa-user-id-v3=s%253AAQAKIE6VkTlryyUUSgqm5ySuauqDmoVCMw8UwTlVJdJVo7VyEHwYAiCZoIvBBjoEVEYQwkIEC_YL7g.TAQw0IhBpAC4cz7DmwBur65gB3YfWUfW9T8HUNZ5zmA; _uetsid=1c372a402f6611f0bb982f950e6cbe1c; _uetvid=232567002e1f11f09ba06fe16b072e04; bm_lso=5BCEDC24BB519919C5B39E3615983A9F9FE94C1449427DCAD0801485020739D9~YAAQPGw/F0t0LYeWAQAASsv8xwOXyhGJElDa++rnq11hVJp+MPvPuWkEeQu5vniwlSLSf1+RDl2rNHpvM7P21/I8cWRGYnZixGz8UfwGvaeswUayUwPJ09DOWzEmtdpfI2+Y/XRrbTA3r/KhuC21MN0bl2iHRnkRB5SNT9nvcK6dIuf23t8naiQp4dVFn5LggHTww0/+Wt5jW+c96HoqhB5WzsEUTvFy02JSqnNuSxE/HXsRK6bfDNVVd8QK2kif+bYkGR+xiDGAqOWmpIN61G2kvaw2Oje+8t/LC+9vM5XbSipZvYN0Z9qLLPaRP77X2Nnvbr2D37q2cK8yEVFPK1inbTr/bU9p9OshHtVre3Osaefz7r4bnFs+sRo858ArHAgGVfuebLSYKyb/DWgV0SL3J6BaIL6WCWlbVSPf6qsoR+qXtRgTt6MnJQwAaVqjWBOtgs6oQTCxsbR+rY/G^1747111962760; da_sid=A8C6EF618E33AECA83D3AA13A291809AE2.1|3|0|3; da_lid=34B14E159A78EA34A2A4BB99E0E9A3C59F|0|0|0; da_intState=; bm_s=YAAQPGw/F2x2LYeWAQAAJw39xwM7gbXmCbTWSuBeSJO/6QFLuVWeq2kJTDYYRcw98FpELJyOyGpwlBXO6Vp5eC/VRO/munk5ixYpHyRGACdN95iQdDxt1ZFwmiaAWuRUVvVd309UxmVRzGa1nRh22AX12Q2qF2wlqht6OJo2EI4jwSI9bP/vWR+fvZFCHarE7EA4wLIQIQt/YkxLgEHqvU9SwbU0hXe+X2FcJIWEDjPfMy9hUBYxbR7kVpK3Iy9GwCJaNNC2MMC6sfo7CwRg7NiatV+FbAEAPLz5Rvn6/51cB0bMAef+T1SdLaMHcGyM0hx+wDxEq8iynxOaGr3BATtd4ly3A6bptkj3yx2ATIChQO8bziHtRTShkCnAlPwijC54WKN2f4mGnAtCvZuFV3VSJxVzR7g9Cy8gXroffi4kfrocIxG+sRsdfS0j0TA4ECfxojx1quQry+SOdxsc8ZHUR6WRN8iCkI7y2GY8mxQlNf8HaH4HLIPz07bLm1RqZTMVh7VbfzevaNohDoFqjngqjPBA1HUznN6YGnfcmDzeOQHMsvzu0UMpbgT0y3tvptCYru/b65I=; _ga_LTF3E70SKQ=GS2.1.s1747111669$o5$g1$t1747112968$j59$l0$h0; bm_sv=7F5398908A050E8FF7F0D74635DD76E0~YAAQXXBWuDuler2WAQAAtEYMyBsRgV/P8Jt3X47vW7BHhAPGM+9FnTsEZy8yPqSEJx2eIQxTT7njVwgrsqDZmfkzQB7DCK68INchIZ1TILPybjCZUVnsOGpfjDOiY2Lpc+F7fy+0thD65aJDWBkIraNhfDfInP35IMkjyIi1sz9ZfprA3Wusi5jnXCUHaRX36H6MoL0rQ+aeecGMWrMXoCkr0t8N2YaCWyO7pU8Gp4Fe7Z1qLtnChKiQ6lCPi432~1',
}


def write_to_excel(data):
    headers = ["Item No.", "Product Name", "Product Description", "Product Images",
               "Pack Sizes", "Overview", "Document Links", "Specifications"]
    filename="products.xlsx"

    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
    else:
        wb = load_workbook(filename)
        ws = wb.active

    ws.append(data)
    wb.save(filename)
    print(f"✅ Data written to {filename}")

def clean_html_description(html_description):
    soup = BeautifulSoup(html_description, "html.parser")

    for sup_tag in soup.find_all("sup"):
        sup_tag.unwrap()

    for b_tag in soup.find_all("b"):
        b_tag.unwrap()

    for ul_tag in soup.find_all("ul"):
        ul_tag.unwrap()
    for li_tag in soup.find_all("li"):
        li_tag.unwrap()

    cleaned_text = soup.get_text("\n", strip=True)

    return cleaned_text

def productDetails(p_url,cookies_dy):
    try:
        time.sleep(2)  

        response = requests.get(p_url
                            ,headers=headers,
                            cookies=cookies_dy)

        # Parse the HTML content with lxml

        tree = html.fromstring(response.content)
        soup = BeautifulSoup(response.text, 'html.parser')

        data = soup.find('script',type='application/ld+json')
        json_data =json.loads(data.string)

        item_no= json_data['sku']
        print("item no ",item_no)
        print("------------------------------------------------------------------------------------")

        product_name = json_data['name']   
        print("product name ",product_name)
        print("------------------------------------------------------------------------------------")

        product_description = clean_html_description(json_data['description'])
        print("product description ",product_description)
        print("------------------------------------------------------------------------------------")

        product_images = json_data['image']
        print("product images ",product_images)
        print("------------------------------------------------------------------------------------")

        # product_overview = driver.find_element(By.XPATH, "//div[@class='bab-tab-pim-content']")
        product_overview = tree.xpath("//div[@class='bab-tab-pim-content']//p | //div[@class='bab-tab-pim-content']//ul")
        product_overview_text = ""

        for el in product_overview:
            if el.tag == 'p':
                # Process <p> tag, get text and add a newline after each <p>
                text = el.text_content().strip() + "\n"
                product_overview_text += text
            elif el.tag == 'ul':
                # Process <ul> tag by handling each <li> item in the <ul>
                product_overview_text += "\n"
                for li in el.xpath(".//li"):
                    li_text = li.text_content().strip()
                    product_overview_text += f"- {li_text}\n"  # Add bullet point style

        print("Product Overview:", product_overview_text)

        print("------------------------------------------------------------------------------------")

        base_url = "https://shop.sartorius.com/medias"
        pdf_links = tree.xpath('//div[@class="product-download__title"]//a[contains(@href, ".pdf")]/@href')
        full_pdf_urls = [base_url + link +" ," if link.startswith('/') else link+" ," for link in pdf_links]
        print(f"✅ Found {len(full_pdf_urls)} PDF link(s).")
        print("------------------------------------------------------------------------------------")
        full_pdf_urls_text = "\n".join(full_pdf_urls)
        print(full_pdf_urls_text)
        print("------------------------------------------------------------------------------------")

        specification_groups = tree.xpath('//div[@class="specifications"]//div[@class="specifications__group"]')
        specification_lines = []

        for group in specification_groups:
            lines = group.xpath(".//text()")
            for line in lines:
                clean_line = line.strip()
                if clean_line:
                    specification_lines.append(clean_line)

        specification_text = "\n".join(specification_lines)
        print(specification_text)
        print("------------------------------------------------------------------------------------")

        print("------------------------------------------------------------------------------------")

        product_price_data = pricedetails.get_product_price(p_url, cookies_dy)
        product_price_data_text = ",\n".join([str(d) for d in product_price_data])
        print(product_price_data_text)

        write_to_excel([
            item_no,
            product_name,
            product_description,
            product_images,
            product_price_data_text,
            product_overview_text,
            full_pdf_urls_text,
            specification_text
        ])


    except Exception as e:
        print(f"An error occurred: {e}")
